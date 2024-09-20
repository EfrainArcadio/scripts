import os
import pandas as pd
import psycopg2
import json
from openpyxl import Workbook
## Primer fecha compuesta de la consulta 
y= "2024"
mes = "Junio"
m = "06"
dia_in = "1"
dia_fn= "30"
###
app = '101801'
dig = '101800'
com = '201A00'
##
## Obtener la ruta del directorio actual
ruta_actual = os.path.dirname(__file__)
## Subir un nivel en el directorio
ruta_superior = os.path.dirname(ruta_actual)
ruta_sup_2 = os.path.dirname(ruta_superior)
##
ruta_dumps = os.path.join(ruta_sup_2,f'public/recargas/meses/{y}/{m} {mes}')
def path_verify(path):
  if not os.path.exists(path):
    os.makedirs(path)
    print(f'Directorio Creado: {path}')
  else:
    print(f'El Directorio ya existe: {path}')
path_verify(ruta_dumps)
###
json_conn = os.path.join(ruta_sup_2,f'config/db.json')
with open(json_conn, 'r') as f_conn:
  data_conn = json.load(f_conn)
connection = psycopg2.connect(**data_conn)
## Consulta para extraer todos los datos del periodo de tiempo seleccionado
cursor = connection.cursor()
cursor.execute(f"SELECT * FROM datos_rre_2024 WHERE fecha_hora_transaccion >= '{y}-{m}-{dia_in} 00:00:00' AND fecha_hora_transaccion <= '{y}-{m}-{dia_fn} 23:59:59'")
transacciones = cursor.fetchall()
##
newDatos = []
for fila in transacciones:
  newDatos.append({
    'id_transaccion_organismo':fila[0],
    'provider': fila[1],
    'tipo_tarjeta': fila[2],
    'numero_serie_hex': fila[3],
    'fecha_hora_transaccion': fila[4],
    'tipo_equipo' : fila[5],
    'location_id' : fila[6],
    'tipo_transaccion' : fila[7],
    'saldo_antes_transaccion' : fila[8],
    'monto_transaccion' : fila[9],
    'saldo_despues_transaccion' : fila[10],
    'sam_serial_hex_ultima_recarga' : fila[11],
    'sam_serial_hex' : fila[12],
    'contador_recargas' : fila[13],
    'event_log' : fila[14],
    'load_log' : fila[15],
    'mac' : fila[16],
    'sam_counter' : fila[17],
    'environment' : fila[18],
    'environmet_issuer_id' : fila[19],
    'contract' : fila[20],
    'contract_tariff' : fila[21],
    'contract_sale_sam' : fila[22],
    'contract_restrict_time' : fila[23],
    'contract_validity_start_date' : fila[24],
    'contract_validity_duration' : fila[25]
  })
##
cursor2 = connection.cursor()
cursor2.execute(f"SELECT * FROM datos_ext_rre_2024 WHERE start_date >= '{y}-{m}-{dia_in} 00:00:00' AND start_date <= '{y}-{m}-{dia_fn} 23:59:59'")
extenciones = cursor2.fetchall()
newExtenciones = []
for fila in extenciones:
  newExtenciones.append({
    'id_transaccion_organismo':fila[0],
    'user_id': fila[1],
    'device_id': fila[2],
    'latitude': fila[3],
    'longitude': fila[4],
    'start_date' : fila[5],
    'end_date' : fila[6],
    'duration' : fila[7],
  })
##
connection.close()
df_transacciones = pd.DataFrame(newDatos)
df_extenciones = pd.DataFrame(newExtenciones)

###############################################################################
#                           Funciones                                         #
###############################################################################
## DataFrame por tipo de red
def df_tipored(df,location_id):
  df_tipo = df[df['location_id'] == location_id]
  return df_tipo
##
def resumen_transacciones(df,fechas_unicas):
  resumen = []  
  ## Analizar dia por dia
  for fecha in fechas_unicas:
    df_dia = df[df['fecha_hora_transaccion'] == fecha]
    ## Crear DataFrame por tipo de red por dia
    df_app = df_tipored(df_dia,app)
    df_dig = df_tipored(df_dia,dig)
    df_com = df_tipored(df_dia,com)
    ## extraer montos por tipo de red
    monto_app = sum(df_app['monto_transaccion']) / 100
    monto_dig = sum(df_dig['monto_transaccion']) / 100
    monto_com = sum(df_com['monto_transaccion']) / 100
    ## Sumar los montos
    monto_total = monto_app + monto_dig + monto_com
    ## Transacciones totales
    tr_totales = len(df_dig) + len(df_app) + len(df_com)  
    ##
    resumen.append({
        'FECHA': fecha,
        'TR Digitales': len(df_dig),
        'TR Fisicas': len(df_com),
        'TR AppCDMX': len(df_app),
        'TR Totales': tr_totales,
        'Montos Digitales': monto_dig,
        'Montos Fisicos': monto_com,
        'Montos AppCDMX': monto_app,
        'Monto Total': monto_total
    })
  return resumen
##
###############################################################################
#                           Funciones                                         #
###############################################################################
##
df_suc = df_transacciones[df_transacciones['tipo_transaccion'] == '0'].copy()
df_suc['fecha_hora_transaccion'] = df_suc['fecha_hora_transaccion'].dt.strftime('%Y-%m-%d')
fechas_unicas = sorted(set(df_suc['fecha_hora_transaccion']))
##

print(df_suc[df_suc['location_id'] == '101800'])

res = resumen_transacciones(df_suc,fechas_unicas)
df_res = pd.DataFrame(res)

df_res_short = df_res[['FECHA','Montos AppCDMX','Montos Digitales','Montos Fisicos','Monto Total']]
##
###
print(f'Creando Reporte de SAM´s {mes}...')
lista_sams = sorted(set(df_transacciones['sam_serial_hex']))
sams = f'Reporte_SAM´s_{mes}.xlsx'
ruta_sams = os.path.join(ruta_dumps,sams)
res = []
with pd.ExcelWriter(ruta_sams) as writer:
  for sam in lista_sams:
    df_sam = df_transacciones[df_transacciones['sam_serial_hex'] == sam]
    acre = len(df_sam[df_sam['tipo_transaccion'] == '0'])
    # print(acre)
    errors = len(df_sam[df_sam['tipo_transaccion'] == 'D0'])
    # print(errors)
    reint = len(df_sam[df_sam['tipo_transaccion'] == '50'])
    # print(reint)
    res.append({
      'SAM': sam,
      'Acreditadas': acre,
      'Errores': errors,
      'Reintentos': reint,
      'Saltos': '',
      'Total':''
    })
    df_res_sams = pd.DataFrame(res) 
    df_sam_ordenado = df_sam[['id_transaccion_organismo', 'contador_recargas']].sort_values(by='contador_recargas', ascending=True)
    df_res_sams.to_excel(writer, index=False ,sheet_name=f'RESUMEN {mes}') 
    df_sam_ordenado.to_excel(writer, index=False ,sheet_name=f'{sam}')
#######
print("Creando Archivo Penalizaciones...")
## Analisis para las penalizaciones de las transacciones mayores a 7 seg
# df_extenciones['duration'] = df_extenciones['duration']
df_7 = df_extenciones[df_extenciones['duration'] > 7].copy()
tr_may_7s = len(df_7)
print('Transacciones > 7S',tr_may_7s)

## Convierte la serie a un tipo de datos numérico
df_short_7 = df_7[['id_transaccion_organismo', 'duration','start_date','end_date']]
## Convierte la serie a un tipo de datos numérico
df_pena = pd.merge(df_short_7, df_transacciones, on='id_transaccion_organismo', how='inner')
df_merge_fil = df_pena[['id_transaccion_organismo', 'location_id', 'monto_transaccion', 'end_date', 'duration']]
df_merge_fil['monto_transaccion'] = df_merge_fil['monto_transaccion'].apply(lambda x: x / 100)

lista_tr_7s = f"RRE - Penalizaciones {mes}.xlsx"
ruta_lista = os.path.join(ruta_dumps,lista_tr_7s)
with pd.ExcelWriter(ruta_lista) as writer:
    df_res_short.to_excel(writer, index=False ,sheet_name=f'{mes} {y}')
    df_merge_fil.to_excel(writer, index=False ,sheet_name=f'Transacciones penalizables {mes}')
###
print(f'Creando Reporte MP {mes}...')
tr_fisicas = sum(df_res['TR Fisicas'])
tr_digitales = sum(df_res['TR Digitales'])
tr_appcdmx = sum(df_res['TR AppCDMX'])
tr_total = sum(df_res['TR Totales'])
##
pr_dig = tr_digitales / tr_total  
pr_fis = tr_fisicas / tr_total  
pr_app = tr_appcdmx / tr_total  
pr_total = pr_dig + pr_fis + pr_app
## Tabla de valores porcentuales para la comision de mercado pago
## Recargas Digitales
## Porcentaje mensual de comision digital
pr_dg1 = 2.2
pr_dg2 = 2.1
pr_dg3 = 2
pr_dg4 = 1.9

## Recargas Fisicas (Negocios)
## Porcentaje mensual de comision fisica
pr_fs1 = 2
pr_fs2 = 1.9
pr_fs3 = 1.8
pr_fs4 = 1.7
## Limites de Rango, esta es la tabla que se ecnuentra en el presente contrato de Mercado Pago
## Los mismos rangos se utilizan para recargas fisicas y digitales 
r1 = 15000000
r2 = 30000000
r3 = 45000000
r4 = 60000000

## Obtener totales en $ para fisicas, digitales y el total de la suma de ambas 
tt_fisico = df_res['Montos Fisicos'].sum()
tt_digital = df_res['Montos Digitales'].sum()
tt_appcdmx = df_res['Montos AppCDMX'].sum()
tt_abs_digital = tt_digital + tt_appcdmx
mt_total = tt_fisico + tt_digital + tt_appcdmx
## Se realizan las condicionales para ajustar el porcentaje automaticamente segun el total segun sea el caso fisica o digital  
## Condicional de porcentaje Digital
if tt_abs_digital <= r1:
    pr_com_dig = pr_dg1
elif tt_abs_digital <= r2:
    pr_com_dig = pr_dg2
elif tt_abs_digital <= r3:
    pr_com_dig = pr_dg3
elif tt_abs_digital <= r4:
    pr_com_dig = pr_dg4
    
## Condicional de porcentaje Fisico
if tt_fisico <= r1:
    pr_com_fis = pr_fs1
elif tt_fisico <= r2:
    pr_com_fis = pr_fs2
elif tt_fisico <= r3:
    pr_com_fis = pr_fs3
elif tt_fisico <= r4:
    pr_com_fis = pr_fs4
    
## Comisiones Fisicas y Digitales
com_fisico = (pr_com_fis/100)*tt_fisico
com_digital = (pr_com_dig/100)*tt_digital
com_appcdmx = (pr_com_dig/100)*tt_appcdmx
com_total = com_fisico + com_digital + com_appcdmx
prm_digital = tt_digital / mt_total 
prm_fisico = tt_fisico / mt_total 
prm_appcdmx = tt_appcdmx / mt_total 
prm_total = prm_digital + prm_fisico + prm_appcdmx

#Resultados Transacciones 
wb = Workbook()

## Si existe una hoja llamada Sheet, se elimina para evitar crear una hoja vacia
if 'Sheet' in wb.sheetnames:
  wb.remove(wb['Sheet'])
    
## Declaramos las hojas en las cuales se van a guardar todas las tablas y guardar la informacion
hoja = wb.create_sheet(title="Reporte Mensual MP")

## Agregar la informacion del total de transacciones y el porcentaje total que ocupan tanto fisicas y digitales
## Se ocupara la hoja 1 para guardar las tablas de montos y transacciones totales con sus porcentajes
hoja['A1'] = "Tabla de Transacciones"
hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
hoja.append(['Cantidad de Recargas',tr_digitales,tr_fisicas,tr_appcdmx,tr_total])
hoja.append(['Proporcion',pr_dig,pr_fis,pr_app,pr_total])

## Crear un campo vacio para darle un espacio entre la primer y segunda tabla
hoja.append([])

## Tabla con la informacion de el total de montos con sus porcentajes para fisicas y digitales en base al rango que ocupen
hoja['A6'] = "Tabla de Montos y Proporciones"
hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
hoja.append(['Cantidad de Recargas',tt_digital,tt_fisico,tt_appcdmx,mt_total])
hoja.append(['Proporcion',prm_digital,prm_fisico,prm_appcdmx,prm_total])
hoja.append(['Comision para MP',com_digital,com_fisico,com_appcdmx,com_total])
## Se guarda el archivo de Mercado pago
archivo_mp = f"Reporte_MP_{mes}.xlsx"
ruta_mp = os.path.join(ruta_dumps ,archivo_mp)
wb.save(ruta_mp)

print("Proceso realizado con Exito!!")
###

