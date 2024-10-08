##importar librearias necesarias para el funcionamiento del script
import os
import pandas as pd
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
## Nombre del mes con texto, se ocupara para leer la carpeta del mes y asignar el nombre a los archivos generados
mes_nombre = "Julio"
## Modificar el contenido de m = "mes" * Para los meses que anteriores a octubre ocupar la sintaxis 09 = Septiembre 08 = Agosto
## Modificar el contenido de Y = "Año" 2023 / 2024 / 2025 
m = "07"
y = "2024"
##
semana = "29"
## Nombre de las extenciones de los archivos que ocupara el script para realizar 
a = "-Transacciones.csv"
ae = "-Transacciones-extension.csv"

## Ruta de la cual se extraeran todos los archivos y en la misma se guardaran los archivos
ruta_guardado = f"data/{y}/{m} {mes_nombre}"

## Este es el rango de dias en el que se trabajara, para el tema del ultimo dia siempre se le sumara 1
## Ejemplo primera quincena dia_fn = 16 el metodo range trabaja de esa forma
dia_in = 15
dia_fn = 22
rango = dia_fn - dia_in

## Listado de los archvios -Transacciones.csv
## Listado de los archivo a leer segun el rango especificado 
archivo_tr = [os.path.join(ruta_guardado, f"{y}{m}{d:02d}{a}") for d in range(dia_in, dia_fn)]
## Leer Archivos de Extencion para obtener la duracion de las transacciones
## Lista de nombres de archivo
archivo_ex = [os.path.join(ruta_guardado, f"{y}{m}{d:02d}{ae}") for d in range(dia_in, dia_fn)]
## Areglo que se llenara con los archivos -Transacciones-extension.csv
transacciones = []
## Bucle para insertar todos los archivos en el DataFrame transacciones
for transaccion in archivo_tr:
  df = pd.read_csv(transaccion, low_memory=False)
  transacciones.append(df)
## Concatenación de documentos extraídos del arreglo transacciones y creando un solo DataFrame con información de toda la quincena
## Arreglo que se llenara con los archivos -Transacciones.csv
extenciones = []
## Bucle para insertar todos los archivos en el DataFrame extenciones
for extencion in archivo_ex:
  df = pd.read_csv(extencion)
  extenciones.append(df)  
## Concatenación de documentos extraídos del arreglo extenciones y creando un solo DataFrame con información de todo el mes

resumen = []

## Funcion para el analisis general
def resumen_transacciones(transacciones):
  for df in transacciones:
    ## Filtrar las transacciones de tipo 0
    df['TIPO_TRANSACCION'] = df['TIPO_TRANSACCION'].astype('str')
    df_filtro = df[df['TIPO_TRANSACCION'] == '0'].copy()
    ## Sacaremos el total de transacciones con el metodo count 
    tr_totales = df_filtro['TIPO_TRANSACCION'].count()
    ## Convertir la columna FECHA_HORA_TRANSACCION a datetime
    df_filtro['FECHA_HORA_TRANSACCION'] = pd.to_datetime(df_filtro['FECHA_HORA_TRANSACCION'])
    df_filtro['FECHA_HORA_TRANSACCION'] = df_filtro['FECHA_HORA_TRANSACCION'].dt.strftime('%Y-%m-%d')
    ## Separar las transacciones en base al método
    df_fisico = df_filtro[df_filtro['LOCATION_ID'] == '201A00']
    df_digital = df_filtro[df_filtro['LOCATION_ID'] == '101800']
    df_appcdmx = df_filtro[df_filtro['LOCATION_ID'] == '101801']
    ## Calcular el monto total por transacción física y agregar al DataFrame correspondiente
    monto_fisico = df_fisico['MONTO_TRANSACCION'].sum()
    ## Calcular el monto total por transacción digital y agregar al DataFrame correspondiente
    monto_digital = df_digital['MONTO_TRANSACCION'].sum()
    ## 
    monto_appcdmx = df_appcdmx['MONTO_TRANSACCION'].sum()
    ## Obtener los valores únicos de la fecha de transacción
    fechas_unicas = df_filtro['FECHA_HORA_TRANSACCION'].unique()
    ## Agregar los resultados a la lista de resumen 
    resumen.append({
        'FECHA': ', '.join(fechas_unicas),
        'TR Digitales': df_digital.shape[0],
        'TR Fisicas': df_fisico.shape[0],
        'TR AppCDMX': df_appcdmx.shape[0],
        'TR Totales': tr_totales,
        'Montos Digitales': monto_digital / 100,
        'Montos Fisicos': monto_fisico / 100,
        'Montos AppCDMX': monto_appcdmx / 100,
        'Monto Total': (monto_digital + monto_fisico + monto_appcdmx)/100,
    })
  return resumen
## Definir funcion para crear graficos:
def grafico(dias,mt,tr,title,name,periodo,colorb,colorl,colort):
  ## Declaracion del Grafico
  plt.style.use('seaborn-v0_8-paper')
  ## Definimos el objeto grafico (ancho,alto)
  fig,ax = plt.subplots(figsize=(18,8))
  ## Definimos el Area que ocupara el grafico
  plt.subplots_adjust(left=0.05, right=0.94, bottom=0.148, top=0.94)
  ## Valores base p
  ax2 = ax.twinx()
  p = ax.bar(dias,tr,label='Transacciones',color=f'#{colorb}')
  for i in enumerate(tr):
    ax.bar_label(p, label_type='center', color='#fff',fontsize=11,**{'fmt': '{:,.0f}'})
  for i, m in enumerate(mt):
    ax2.annotate(f'${m:,.0f}', (dias[i], m), xytext=(-20,10), textcoords='offset points', fontsize=12, fontweight=600,color=f'#{colort}')  
  ax2.plot(dias,mt,label='Montos',color=f'#{colorl}', marker='o',linestyle='solid')
  ## Creamos una legenda fuera del grafico
  fig.legend(loc='lower center', ncols=2, fontsize=12)
  # Ajusta el formato de los valores en el eje Y
  ax.yaxis.set_major_formatter(lambda x, pos: f'{x:,.0f}')
  ax2.yaxis.set_label_position("right")
  ax2.yaxis.set_major_formatter(lambda x, pos: f'${x:,.0f}')
  ##
  ax.tick_params(axis='x', labelsize=10)
  ax.tick_params(axis='y', labelsize=10)
  ax2.tick_params(axis='y', labelsize=10)
  ax.set_title(title,fontsize=14,fontweight=600)
  ax.set_xlabel(periodo,fontsize=12,fontweight=600)
  ax.set_ylabel('Transacciones',fontsize=12,fontweight=600)
  ax2.set_ylabel('Montos',fontsize=12,fontweight=600)
  ####################### Guarda el gráfico en alta resolución
  ruta_grafico = os.path.join(ruta_guardado, name)
  plt.savefig(ruta_grafico,format='png',dpi=980,bbox_inches='tight')

def grafico_rre(df,title,name,periodo):
  ## Grafico
  dias = df['FECHA']
  digitales_tr = df['TR Digitales']  
  appcdmx_tr = df['TR AppCDMX']
  comercios_tr = df['TR Fisicas']
  digitales_mt = df['Montos Digitales']  
  appcdmx_mt = df['Montos AppCDMX']
  comercios_mt = df['Montos Fisicos']
    
    ## Definir el estilo del grafico
  plt.style.use('seaborn-v0_8-paper')
    ## Definimos el objeto grafico (ancho,alto)
  fig,ax = plt.subplots(figsize=(18,8))
  plt.subplots_adjust(left=0.05, right=0.926, bottom=0.148, top=0.94)
    ## Valores base para las barras apiladas
    ## Colores para las barras
# Create stacked bars using bar_stack
  # Create a dataframe to simplify plotting
  tipos = {
      'AppCDMX': appcdmx_mt,
      'Comercios': comercios_mt,
      'Digitales': digitales_mt
  }

  x = np.arange(len(dias))  # the label locations
  width = 0.25  # the width of the bars
  multiplier = 0

  colors = {  # Dictionary of colors for each attribute
      'AppCDMX': '#b81532',
      'Comercios': '#af38c1',
      'Digitales': '#08acec'
  }
  lColors = {
    'AppCDMX': '#4b0615',
    'Comercios': '#420c46',
    'Digitales': '#06314b'
  }
  for attribute, measurement in tipos.items():
      offset = width * multiplier
      color = colors.get(attribute)  # Get color from dictionary
      lcolor = lColors.get(attribute)  
      rects = ax.bar(x + offset, measurement, width, label=attribute, color=color)
      ax.bar_label(rects,label_type='center',padding=2,color=lcolor,fontweight=600,fontsize=9, labels=[f'${value:,.0f}' for value in measurement])
      multiplier += 1
  ax2 = ax.twinx()
      
  for i, tr in enumerate(digitales_tr):
    ax2.annotate(f'{tr:,.0f}', (dias[i], tr), xytext=(-20,10), textcoords='offset points', fontsize=9, color='#06314b')
  ax2.plot(dias,digitales_tr,label='Transacciones Digitales',color='#006fa5', marker='o',linestyle='solid')
  
  for i, tr in enumerate(appcdmx_tr):
    ax2.annotate(f'{tr:,.0f}', (dias[i], tr), xytext=(-20,10), textcoords='offset points', fontsize=9, color='#4b0615')
  ax2.plot(dias,appcdmx_tr,label='Transacciones AppCDMX',color='#f8747e', marker='o',linestyle='solid')
  
  for i, tr in enumerate(comercios_tr):
    ax2.annotate(f'{tr:,.0f}', (dias[i], tr), xytext=(-20,10), textcoords='offset points', fontsize=9, color='#420c46')
  ax2.plot(dias,comercios_tr,label='Transacciones Comercios',color='#66246b', marker='o',linestyle='solid')
  
    ## Creamos una legenda fuera del grafico
  fig.legend(loc='lower center', ncols=6, fontsize=12)
    # Ajusta el formato de los valores en el eje Y
  ax2.yaxis.set_label_position("right")
  ax.yaxis.set_major_formatter(lambda x, pos: f'${x:,.0f}')
  ax2.yaxis.set_major_formatter(lambda x, pos: f'{x:,.0f}')
    # Ajusta el título del gráfico
  ax.tick_params(axis='x', labelsize=10)
  ax.tick_params(axis='y', labelsize=10)
  ax2.tick_params(axis='y', labelsize=10)
  ax.set_title(title,fontsize=14,fontweight=600)
  ax.set_xlabel(periodo,fontsize=12,fontweight=600)
  ax.set_ylabel('Transacciones',fontsize=12,fontweight=600)
  ax2.set_ylabel('Montos',fontsize=12,fontweight=600)
    # Guarda el gráfico en alta resolución
  ruta_grafico = os.path.join(ruta_guardado, name)
  plt.savefig(ruta_grafico,format='png',dpi=980,bbox_inches='tight')

## Crear Resumen del Periodo
resumen_transacciones(transacciones)
resultados = pd.DataFrame(resumen)
## Convertir el arreglo resumen en DataFrame

df_extenciones = pd.concat(extenciones, ignore_index=True)
df_transacciones = pd.concat(transacciones, ignore_index=True)
## Inicia el condicional para los dias sueltos
if rango == 7 :
  print(f"Trabajando con la semana {semana}")
  print("Generando datos en crudo")
  archivo_full = f"Full_semana_{semana}_{mes_nombre}.csv"
  ruta_full = os.path.join(ruta_guardado, archivo_full)
  df_transacciones.to_csv(ruta_full, index=False)
  ##   Extenciones
  print("Generando datos extenciones")
  archivo_full = f"Full_ext_semana_{semana}_{mes_nombre}.csv"
  ruta_full = os.path.join(ruta_guardado, archivo_full)
  df_extenciones.to_csv(ruta_full, index=False)

  ###### Graficos
  ## AppCDMX
  print(f"Creando Grafico AppCDMX..")
  grafico( resultados['FECHA'], resultados['Montos AppCDMX'], resultados['TR AppCDMX'],
    f'Comportamiento AppCDMX semana {semana}', f'AppCDMX_Grafico_Semana_{semana}.png', f'Semana {semana}',
    'b81532', 'f8747e', '4b0615' )
  ## Digital
  print(f"Creando Grafico Digitales..")
  grafico( resultados['FECHA'], resultados['Montos Digitales'], resultados['TR Digitales'],
    f'Comportamiento Digitales semana {semana}', f'Dig_Grafico_Semana_{semana}.png', f'Semana {semana}',
    '08acec', '006fa5', '06314b' )
  ## Comercios
  print(f"Creando Grafico Comercios..")
  grafico( resultados['FECHA'], resultados['Montos Fisicos'], resultados['TR Fisicas'],
    f'Comportamiento Comercios semana {semana}', f'Com_Grafico_Semana_{semana}.png', f'Semana {semana}',
    'af38c1', '66246b', '420c46' )
  ##
  nameGrRRE = f'RRE_Grafico_Semana_{semana}.png'
  titleRRE = f'Comportamiento RRE semana {semana}'
  periodoRRE = f'Semana {semana}'
  grafico_rre(resultados,titleRRE,nameGrRRE,periodoRRE)
  print('Proceso Finalizado con Exito!!')
  ## 
  if rango < 13 :
    print("Generando datos en crudo")

    archivo_full = f"Full_{dia_in}-{dia_fn-1}_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    
    archivo_sem = f"RRE_{mes_nombre}_{dia_in}-{dia_fn-1}.csv"
    ruta_res_sem = os.path.join(ruta_guardado, archivo_sem)
    resultados.to_csv(ruta_res_sem, index=False)
    print("Proceso realizado con Exito!!")
## Inicia el condicional para las Quincenas        
elif rango >= 13 and rango <= 16:
  print("Realizando analisis quincenal.")
  ## Nombres de las quincenas
  first = "1ra"
  second = "2da"
 
  ## Arreglo para almacenar el resumen de las tarjetas
  tarjetas = []

  # Filtrar las transacciones de tipo 0
  df_filtered = df_transacciones[df_transacciones['TIPO_TRANSACCION'] == '0'].copy()
  df_appcdmx = df_filtered[df_filtered['LOCATION_ID'] == '101801']    
  df_digital = df_filtered[df_filtered['LOCATION_ID'] == '101800']    
  df_fisicas = df_filtered[df_filtered['LOCATION_ID'] == '201A00']    
  # Obtener el monto total
  montoAppcdmx = df_appcdmx['MONTO_TRANSACCION'] / 100
  montoDigital = df_digital['MONTO_TRANSACCION'] / 100
  montoFisico = df_fisicas['MONTO_TRANSACCION'] / 100
  
  # Obtener las tarjetas únicas
  tarjetas_unicas = df_filtered['NUMERO_SERIE_HEX'].unique()

  # Agregar los resultados a la lista de tarjetas
  tarjetas.append({
    '# Tarjetas': len(tarjetas_unicas),
    '# Transacciones': len(df_filtered),
    '$ Promedio Digital': montoDigital.mean(),
    '$ Promedio Fisico': montoFisico.mean(),
    '$ Promedio AppCDMX': montoAppcdmx.mean(),
  })

  if dia_in == 1:
    ## Creando RRE
    print(f"Analizando la {first} qna de {mes_nombre}")
    ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
    print(f"Generando RRE {first} qna {mes_nombre}")
    archivo_mens = f"RRE_{first}_qna_{mes_nombre}.csv"
    ruta_res_mens = os.path.join(ruta_guardado, archivo_mens)
    resultados.to_csv(ruta_res_mens, index=False)
    ## Archivo completo
    print("Generando datos en crudo")
    archivo_full = f"Full_{first}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    ##   Ectenciones
    print("Generando datos extenciones")
    archivo_full = f"Full_ext_{first}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_extenciones.to_csv(ruta_full, index=False)
    ## Resumen de Tarjetas
    print("Generando resumen de las tarjetas")
    res_tarjetas = pd.DataFrame(tarjetas)
    archivo_tar = f"Tarjetas_{first}_qna_{mes_nombre}.csv"
    ruta_resultados = os.path.join(ruta_guardado, archivo_tar)
    res_tarjetas.to_csv(ruta_resultados, index=False)
    
    ## AppCDMX
    print(f"Creando Grafico AppCDMX..")
    grafico( resultados['FECHA'], resultados['Montos AppCDMX'], resultados['TR AppCDMX'],
      f'Comportamiento AppCDMX {first} qna {mes_nombre}', f'AppCDMX_Grafico_{first}_qna_{mes_nombre}.png', f'{first} qna {mes_nombre}',
      'b81532', 'f8747e', '4b0615' )
    ## Digital
    print(f"Creando Grafico Digitales..")
    grafico(resultados['FECHA'], resultados['Montos Digitales'], resultados['TR Digitales'],
      f'Comportamiento Digitales {first} qna {mes_nombre}', f'Dig_Grafico_{first}_qna_{mes_nombre}.png', f'{first} qna {mes_nombre}'
      '08acec', '006fa5', '06314b' )
    ## Comercios
    print(f"Creando Grafico Comercios..")
    grafico(resultados['FECHA'], resultados['Montos Fisicos'], resultados['TR Fisicas'], 
            f'Comportamiento Comercios {first} qna {mes_nombre}', f'Com_Grafico_{first}_qna_{mes_nombre}.png', f'{first} qna {mes_nombre}', 
            'af38c1', '66246b', '420c46' )
    ##
    nameGrRRE = f'RRE_Grafico_{first}_qna_{mes_nombre}.png'
    titleRRE = f'Comportamiento RRE {first} qna {mes_nombre}'
    periodoRRE = f'{first} qna {mes_nombre}'
    grafico_rre(resultados,titleRRE,nameGrRRE,periodoRRE)
  ## Condicion para la segunda Quincena
  elif dia_in == 16:
    print(f"Analizando la {second} qna de {mes_nombre}")
    ## Creando RRE
    print(f"Generando RRE {second} qna {mes_nombre}")
    ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
    archivo_mens = f"RRE_{second}_qna_{mes_nombre}.csv"
    ruta_res_mens = os.path.join(ruta_guardado, archivo_mens)
    resultados.to_csv(ruta_res_mens, index=False)
    ## Archivo completo 
    print("Generando datos en crudo")
    archivo_full = f"Full_{second}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_transacciones.to_csv(ruta_full, index=False)
    ## Extenciones    
    print("Generando datos extenciones")
    archivo_full = f"Full_ext_{second}_qna_{mes_nombre}.csv"
    ruta_full = os.path.join(ruta_guardado, archivo_full)
    df_extenciones.to_csv(ruta_full, index=False)
    ## Resumen de tarjetas
    print("Generando resumen de las tarjetas")
    res_tarjetas = pd.DataFrame(tarjetas)
    archivo_tar = f"Tarjetas_{second}_qna_{mes_nombre}.csv"
    ruta_resultados = os.path.join(ruta_guardado, archivo_tar)
    res_tarjetas.to_csv(ruta_resultados, index=False)
    ## Grafico
    ## AppCDMX
    print(f"Creando Grafico AppCDMX..")
    grafico( resultados['FECHA'], resultados['Montos AppCDMX'], resultados['TR AppCDMX'],
      f'Comportamiento AppCDMX {second} qna {mes_nombre}', f'AppCDMX_Grafico_{second}_qna_{mes_nombre}.png', f'{second} qna {mes_nombre}',
      'b81532', 'f8747e', '4b0615' )
    ## Digital
    print(f"Creando Grafico Digitales..")
    grafico(resultados['FECHA'], resultados['Montos Digitales'], resultados['TR Digitales'],
      f'Comportamiento Digitales {second} qna {mes_nombre}', f'Dig_Grafico_{second}_qna_{mes_nombre}.png', f'{second} qna {mes_nombre}'
      '08acec', '006fa5', '06314b' )
    ## Comercios
    print(f"Creando Grafico Comercios..")
    grafico(resultados['FECHA'], resultados['Montos Fisicos'], resultados['TR Fisicas'], 
            f'Comportamiento Comercios {second} qna {mes_nombre}', f'Com_Grafico_{second}_qna_{mes_nombre}.png', f'{second} qna {mes_nombre}', 
            'af38c1', '66246b', '420c46' )
    ##
    nameGrRRE = f'RRE_Grafico_{second}_qna_{mes_nombre}.png'
    titleRRE = f'Comportamiento RRE {second} qna {mes_nombre}'
    periodoRRE = f'{second} qna {mes_nombre}'
    grafico_rre(resultados,titleRRE,nameGrRRE,periodoRRE)
## Inicia el Condicional para los meses    
elif rango > 16:
  ## Se guarda la concatenacion de todo el mes para conciliacion con SEMOVI
  print("Generando datos en crudo")
  archivo_full = f"Full_{mes_nombre}.csv"
  ruta_full = os.path.join(ruta_guardado, archivo_full)
  df_transacciones.to_csv(ruta_full, index=False)
  ## Extenciones    
  print("Generando datos extenciones")
  archivo_full = f"Full_ext_{mes_nombre}.csv"
  ruta_full = os.path.join(ruta_guardado, archivo_full)
  df_extenciones.to_csv(ruta_full, index=False)
  ## Definir el nombre de los archivos que seran guardados en la carpeta al finalizar el analisis
  archivo_mens = f"RRE_{mes_nombre}.csv"
  ruta_res_mens = os.path.join(ruta_guardado, archivo_mens)
  resultados.to_csv(ruta_res_mens, index=False)
  ## Aqui inicia el proceso para el analisis del documento de mercado pago
  ########################################################################
  ## Totales de transacciones fisicas para poder realizar operaciones
  tr_fisicas = resultados['TR Fisicas'].sum()
  tr_digitales = resultados['TR Digitales'].sum()
  tr_appcdmx = resultados['TR AppCDMX'].sum()
  tr_total = resultados['TR Totales'].sum()

  ## Obtener datos porcentuales de transacciones fisicas, digitales y el total
  pr_dig = tr_digitales / tr_total  
  pr_fis = tr_fisicas / tr_total  
  pr_app = tr_appcdmx / tr_total  
  pr_total = pr_dig + pr_fis + pr_app

  ## Tabla de valores porcentuales para la comision de mercado pago
  ## Recargas Digitales
  ## Porcentaje mensual de comision digital
  pr_dg1 = 2.2
  pr_dg2 = 2.1
  pr_dg3 = 2
  pr_dg4 = 1.9

  ## Recargas Fisicas (Negocios)
  ## Porcentaje mensual de comision fisica
  pr_fs1 = 2
  pr_fs2 = 1.9
  pr_fs3 = 1.8
  pr_fs4 = 1.7

  ## Limites de Rango, esta es la tabla que se ecnuentra en el presente contrato de Mercado Pago
  ## Los mismos rangos se utilizan para recargas fisicas y digitales 
  r1 = 15000000
  r2 = 30000000
  r3 = 45000000
  r4 = 60000000

  ## Obtener totales en $ para fisicas, digitales y el total de la suma de ambas 
  tt_fisico = resultados['Montos Fisicos'].sum()
  tt_digital = resultados['Montos Digitales'].sum()
  tt_appcdmx = resultados['Montos AppCDMX'].sum()
  tt_abs_digital = tt_digital + tt_appcdmx
  print(tt_abs_digital)
  mt_total = tt_fisico + tt_digital + tt_appcdmx

  ## Se realizan las condicionales para ajustar el porcentaje automaticamente segun el total segun sea el caso fisica o digital  
  ## Condicional de porcentaje Digital
  if tt_abs_digital <= r1:
      pr_com_dig = pr_dg1
  elif tt_abs_digital <= r2:
      pr_com_dig = pr_dg2
  elif tt_abs_digital <= r3:
      pr_com_dig = pr_dg3
  elif tt_abs_digital <= r4:
      pr_com_dig = pr_dg4
      
  ## Condicional de porcentaje Fisico
  if tt_fisico <= r1:
      pr_com_fis = pr_fs1
  elif tt_fisico <= r2:
      pr_com_fis = pr_fs2
  elif tt_fisico <= r3:
      pr_com_fis = pr_fs3
  elif tt_fisico <= r4:
      pr_com_fis = pr_fs4
      
  ## Comisiones Fisicas y Digitales
  com_fisico = (pr_com_fis/100)*tt_fisico
  com_digital = (pr_com_dig/100)*tt_digital
  com_appcdmx = (pr_com_dig/100)*tt_appcdmx
  com_total = com_fisico + com_digital + com_appcdmx
  prm_digital = tt_digital / mt_total 
  prm_fisico = tt_fisico / mt_total 
  prm_appcdmx = tt_appcdmx / mt_total 
  prm_total = prm_digital + prm_fisico + prm_appcdmx
  
  #Resultados Transacciones 
  wb = Workbook()

  ## Si existe una hoja llamada Sheet, se elimina para evitar crear una hoja vacia
  if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
      
  ## Declaramos las hojas en las cuales se van a guardar todas las tablas y guardar la informacion
  hoja = wb.create_sheet(title="Reporte Mensual MP")

  ## Agregar la informacion del total de transacciones y el porcentaje total que ocupan tanto fisicas y digitales
  ## Se ocupara la hoja 1 para guardar las tablas de montos y transacciones totales con sus porcentajes
  hoja['A1'] = "Tabla de Transacciones"
  hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
  hoja.append(['Cantidad de Recargas',tr_digitales,tr_fisicas,tr_appcdmx,tr_total])
  hoja.append(['Proporcion',pr_dig,pr_fis,pr_app,pr_total])

  ## Crear un campo vacio para darle un espacio entre la primer y segunda tabla
  hoja.append([])

  ## Tabla con la informacion de el total de montos con sus porcentajes para fisicas y digitales en base al rango que ocupen
  hoja['A6'] = "Tabla de Montos y Proporciones"
  hoja.append(['Tipo de Recarga','Digital','Fisica','AppCDMX','Total'])
  hoja.append(['Cantidad de Recargas',tt_digital,tt_fisico,tt_appcdmx,mt_total])
  hoja.append(['Proporcion',prm_digital,prm_fisico,prm_appcdmx,prm_total])
  hoja.append(['Comision para MP',com_digital,com_fisico,com_appcdmx,com_total])
  ## Se guarda el archivo de Mercado pago
  archivo_mp = f"Reporte_MP_{mes_nombre}.xlsx"
  ruta_mp = os.path.join(ruta_guardado,archivo_mp)
  wb.save(ruta_mp)
  ########################################################################
  ## Analisis para las penalizaciones de las transacciones mayores a 7 seg
  mayor_seven = df_extenciones['DURATION'] > 7
  
  ## Se realiza el conteo total de todas las transacciones que son mayores a los 7 segundos
  ntr_may_seven = df_extenciones.loc[mayor_seven,['DURATION']].count()
  
  ## Convierte la serie a un tipo de datos numérico
  list_mayor_seven = df_extenciones.loc[mayor_seven,['ID_TRANSACCION_ORGANISMO', 'DURATION','END_DATE']]
  
  ## Convierte la serie a un tipo de datos numérico
  df_merge_succ = pd.merge(list_mayor_seven, df_transacciones, on='ID_TRANSACCION_ORGANISMO', how='inner')
  df_merge_fil = df_merge_succ[['ID_TRANSACCION_ORGANISMO', 'LOCATION_ID', 'MONTO_TRANSACCION', 'END_DATE', 'DURATION']]
  df_merge_fil['MONTO_TRANSACCION'] = df_merge_fil['MONTO_TRANSACCION'].apply(lambda x: x / 100)

  lista_tr_7s = f"RRE - Penalizaciones {mes_nombre}.xlsx"
  ruta_lista = os.path.join(ruta_guardado,lista_tr_7s)
  with pd.ExcelWriter(ruta_lista) as writer:
      df_merge_fil.to_excel(writer, index=False ,sheet_name=f'Transacciones penalizables {mes_nombre}')

  print("Proceso realizado con Exito!!")